import os
import re
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputMediaPhoto
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ================= TOKEN & GROUPS =================
API_TOKEN = os.getenv("BOT_TOKEN")

if not API_TOKEN:
    raise ValueError("❌ BOT_TOKEN environment variable not set!")

GROUP_NEW = -1003735668749
GROUP_DESIGN = -1003867470006
GROUP_READY = -1003312397488
GROUP_SENT = -1003671523271
GROUP_ISSUES = -1003747379674

GROUPS_MAP = {
    "new": GROUP_NEW,
    "design": GROUP_DESIGN,
    "ready": GROUP_READY,
    "sent": GROUP_SENT,
    "issues": GROUP_ISSUES
}

GROUPS_NAMES = {
    GROUP_NEW: "طلبات جديدة",
    GROUP_DESIGN: "تم التصميم",
    GROUP_READY: "مجهز",
    GROUP_SENT: "تم الإرسال",
    GROUP_ISSUES: "مشاكل"
}

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
orders_data = {}

# ================= STATES =================
class OrderState(StatesGroup):
    name = State()
    phone = State()
    city = State()
    area = State()
    order_type = State()
    pieces = State()
    over_type = State()
    hand_type = State()
    box_color = State()
    dist_count = State()
    size = State()
    price = State()
    notes = State()
    images = State()

# ================= EXCEL FUNCTIONS =================
def init_excel_file(file_name: str = "orders.xlsx"):
    """إنشاء ملف Excel إذا كان غير موجود"""
    try:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            ws.append([
                "رقم الطلب",
                "الاسم",
                "الهاتف",
                "المدينة",
                "المنطقة",
                "النوع",
                "القطع",
                "الأوفر",
                "الملحف",
                "لون البوكس",
                "عدد التوزيعات",
                "القياس",
                "السعر",
                "ملاحظات",
                "التاريخ"
            ])
            wb.save(file_name)
            print(f"✅ تم إنشاء الملف: {file_name}")
        else:
            print(f"✅ الملف موجود: {file_name}")
    except Exception as e:
        print(f"❌ خطأ في إنشاء الملف: {e}")

def get_next_order_id(file_name: str = "orders.xlsx"):
    """احصل على رقم الطلب التالي"""
    try:
        if not os.path.exists(file_name):
            return 1
        
        wb = load_workbook(file_name)
        ws = wb.active
        ids = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value and isinstance(row[0].value, int):
                ids.append(row[0].value)
        
        return max(ids) + 1 if ids else 1
    except Exception as e:
        print(f"❌ خطأ في قراءة الإكسل: {e}")
        return 1

def save_to_excel(data, file_name: str = "orders.xlsx"):
    """احفظ الطلب في ملف Excel"""
    try:
        init_excel_file(file_name)
        wb = load_workbook(file_name)
        ws = wb.active
        
        ws.append([
            data.get("id"),
            data.get("name"),
            data.get("phone"),
            data.get("city"),
            data.get("area"),
            data.get("order_type"),
            ",".join(data.get("pieces", [])),
            data.get("over_type", "لا يوجد"),
            data.get("hand_type", "لا يوجد"),
            data.get("box_color", "لا يوجد"),
            data.get("dist_count", "لا يوجد"),
            data.get("size"),
            data.get("price"),
            data.get("notes"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        
        wb.save(file_name)
        print(f"✅ تم حفظ الطلب #{data['id']} في {file_name}")
        
    except Exception as e:
        print(f"❌ خطأ في حفظ الإكسل: {e}")

def create_ready_orders_file():
    """إنشاء ملف بالطلبات الموجودة في كروب مجهز فقط"""
    try:
        ready_file = "orders_ready_current.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Ready Orders"
        
        # أضف رؤوس الأعمدة
        ws.append([
            "رقم الطلب",
            "الاسم",
            "الهاتف",
            "المدينة",
            "المنطقة",
            "النوع",
            "القطع",
            "الأوفر",
            "الملحف",
            "لون البوكس",
            "عدد التوزيعات",
            "القياس",
            "السعر",
            "ملاحظات"
        ])
        
        # أضف فقط الطلبات في كروب مجهز
        for order_id, order_info in orders_data.items():
            if order_info.get("current_group") == "ready":
                data = order_info.get("data", {})
                ws.append([
                    order_id,
                    data.get("name"),
                    data.get("phone"),
                    data.get("city"),
                    data.get("area"),
                    data.get("order_type"),
                    ",".join(data.get("pieces", [])),
                    data.get("over_type", "لا يوجد"),
                    data.get("hand_type", "لا يوجد"),
                    data.get("box_color", "لا يوجد"),
                    data.get("dist_count", "لا يوجد"),
                    data.get("size"),
                    data.get("price"),
                    data.get("notes")
                ])
        
        wb.save(ready_file)
        print(f"✅ تم إنشاء ملف الطلبات الجاهزة: {ready_file}")
        return ready_file
    
    except Exception as e:
        print(f"❌ خطأ في إنشاء ملف الجاهزة: {e}")
        return None

# ================= VALIDATION FUNCTIONS =================
def validate_phone(phone: str) -> bool:
    phone = phone.strip()
    return bool(re.match(r'^[0-9\+\-\s\(\)]{7,15}$', phone))

def validate_price(price: str) -> bool:
    price = price.strip()
    try:
        float(price)
        return True
    except:
        return False

def validate_dist_count(count: str) -> bool:
    count = count.strip()
    try:
        return int(count) > 0
    except:
        return False

# ================= HELPER FUNCTIONS =================
def get_status_buttons(order_id: int, current_group: str = "new") -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    
    if current_group != "new":
        kb.insert(InlineKeyboardButton("⬅️ طلبات جديدة", callback_data=f"move_{order_id}_new"))
    
    if current_group != "design":
        kb.insert(InlineKeyboardButton("✏️ تم التصميم", callback_data=f"move_{order_id}_design"))
    
    if current_group != "ready":
        kb.insert(InlineKeyboardButton("📦 مجهز", callback_data=f"move_{order_id}_ready"))
    
    if current_group != "sent":
        kb.insert(InlineKeyboardButton("✈️ تم الإرسال", callback_data=f"move_{order_id}_sent"))
    
    if current_group != "issues":
        kb.insert(InlineKeyboardButton("⚠️ مشاكل", callback_data=f"move_{order_id}_issues"))
    
    return kb

def format_order_text(data: dict, order_id: int, current_group: str = "new") -> str:
    over = data.get("over_type", "لا يوجد")
    hand = data.get("hand_type", "لا يوجد")
    box = data.get("box_color", "لا يوجد")
    dist = data.get("dist_count", "لا يوجد")
    group_display = GROUPS_NAMES.get(GROUPS_MAP.get(current_group), "غير معروف")

    text = f"""📦 *طلب #{order_id}*

👤 *الاسم:* {data['name']}
📞 *الهاتف:* {data['phone']}
📍 *المدينة - المنطقة:* {data['city']} - {data['area']}

🧵 *النوع:* {data['order_type']}
👕 *القطع:* {', '.join(data['pieces'])}

👗 *الأوفر:* {over}
🛏 *الملحف:* {hand}
🎁 *لون البوكس:* {box}
🎉 *عدد التوزيعات:* {dist}

📏 *القياس:* {data['size']}
💰 *السعر:* {data['price']} ر.س

📝 *الملاحظات:*
{data['notes']}

━━━━━━━━━━━━━━━━━━
📍 *الحالة الحالية:* {group_display}"""
    return text

# ================= KEYBOARDS =================
# قائمة المحافظات العراقية
cities_list = [
    "بغداد",
    "الناصرية - ذي قار",
    "ديالى",
    "الكوت - واسط",
    "كربلاء",
    "دهوك",
    "بابل - الحلة",
    "النجف",
    "البصرة",
    "اربيل",
    "كركوك",
    "السليمانية",
    "صلاح الدين",
    "الانبار",
    "السماوة - المثنى",
    "الموصل",
    "الديوانية",
    "العمارة - ميسان"
]

def get_cities_kb() -> InlineKeyboardMarkup:
    """لوحة مفاتيح المحافظات"""
    kb = InlineKeyboardMarkup(row_width=2)
    for city in cities_list:
        kb.insert(InlineKeyboardButton(f"📍 {city}", callback_data=f"city_{city}"))
    return kb

def get_order_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(
        InlineKeyboardButton("🖨 طباعة", callback_data="type_print"),
        InlineKeyboardButton("🧵 تطريز", callback_data="type_emb")
    )
    return kb

pieces_list = [
    "سيت 3", "سيت 6", "أوفر", "كلو", "صدرية", 
    "حضينة وكماط", "ملحف", "بوكس ككو", "توزيعات"
]

def get_pieces_kb(selected: list) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for p in pieces_list:
        mark = "✅" if p in selected else "☐"
        kb.insert(InlineKeyboardButton(f"{mark} {p}", callback_data=f"piece_{p}"))
    kb.add(InlineKeyboardButton("✔️ تم", callback_data="done_pieces"))
    return kb

def get_over_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🎀 دانتيل", callback_data="over_دانتيل"),
        InlineKeyboardButton("🧵 طباكات", callback_data="over_طباكات"),
        InlineKeyboardButton("📄 صفح", callback_data="over_صفح"),
        InlineKeyboardButton("🎀🧵 دانتيل+طباكات", callback_data="over_دانتيل+طباكات")
    )
    return kb

def get_hand_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🎀 كشكش", callback_data="hand_كشكش"),
        InlineKeyboardButton("🌸 حب الرمان", callback_data="hand_حب الرمان")
    )
    return kb

def get_box_color_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("⚪ أبيض", callback_data="box_أبيض"),
        InlineKeyboardButton("⚫ رصاصي", callback_data="box_رصاصي"),
        InlineKeyboardButton("🩷 وردي", callback_data="box_وردي"),
        InlineKeyboardButton("🩵 سماوي", callback_data="box_سماوي")
    )
    return kb

sizes = ["حديثي ولادة", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]

def get_size_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=4)
    for s in sizes:
        kb.insert(InlineKeyboardButton(s, callback_data=f"size_{s}"))
    return kb

# ================= HANDLERS =================
@dp.message_handler(commands=['start'])
async def cmd_start(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👋 مرحباً!\n\n/new - إنشاء طلب جديد\n/download - تحميل ملف الطلبات الجاهزة")

@dp.message_handler(commands=['new'])
async def cmd_new(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👤 اسم الزبون:")
    await OrderState.name.set()

@dp.message_handler(commands=['download'])
async def cmd_download(msg: types.Message):
    """تحميل ملف الطلبات الموجودة في كروب مجهز فقط"""
    
    # تحقق من وجود طلبات في كروب مجهز
    ready_orders = {oid: info for oid, info in orders_data.items() 
                    if info.get("current_group") == "ready"}
    
    if not ready_orders:
        await msg.answer("❌ لا توجد طلبات في كروب 'مجهز' حتى الآن!")
        return
    
    try:
        # أنشئ ملف Excel بالطلبات الجاهزة فقط
        file_path = create_ready_orders_file()
        
        if file_path and os.path.exists(file_path):
            with open(file_path, 'rb') as file:
                await bot.send_document(
                    chat_id=msg.from_user.id,
                    document=types.InputFile(file_path),
                    caption=f"📊 ملف الطلبات الجاهزة\n\n📦 عدد الطلبات: {len(ready_orders)}"
                )
            await msg.answer("✅ تم إرسال الملف!")
        else:
            await msg.answer("❌ حدث خطأ في إنشاء الملف!")
    
    except Exception as e:
        print(f"❌ خطأ في التحميل: {e}")
        await msg.answer(f"�� خطأ: {str(e)}")

@dp.message_handler(state=OrderState.name)
async def process_name(msg: types.Message, state: FSMContext):
    name = msg.text.strip()
    if len(name) < 2:
        await msg.answer("❌ الاسم قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(name=name)
    await msg.answer("📞 رقم الهاتف:")
    await OrderState.phone.set()

@dp.message_handler(state=OrderState.phone)
async def process_phone(msg: types.Message, state: FSMContext):
    phone = msg.text.strip()
    if not validate_phone(phone):
        await msg.answer("❌ صيغة الهاتف غير صحيحة، حاول مرة أخرى:")
        return
    await state.update_data(phone=phone)
    await msg.answer("📍 المدينة:")
    await OrderState.city.set()

@dp.message_handler(state=OrderState.city)
async def process_city(msg: types.Message, state: FSMContext):
    city = msg.text.strip()
    if len(city) < 2:
        await msg.answer("❌ اسم المدينة قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(city=city)
    await msg.answer("🏘 اختر المحافظة:", reply_markup=get_cities_kb())
    await OrderState.area.set()

@dp.callback_query_handler(lambda c: c.data.startswith("city_"), state=OrderState.area)
async def process_area(call: types.CallbackQuery, state: FSMContext):
    area = call.data.replace("city_", "")
    await state.update_data(area=area)
    await call.message.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
    await OrderState.order_type.set()

@dp.callback_query_handler(lambda c: c.data.startswith("type_"), state=OrderState.order_type)
async def process_order_type(call: types.CallbackQuery, state: FSMContext):
    order_type = "طباعة" if call.data == "type_print" else "تطريز"
    await state.update_data(order_type=order_type)
    await call.message.edit_text("👕 اختر القطع:", reply_markup=get_pieces_kb([]))
    await state.update_data(pieces=[])
    await OrderState.pieces.set()

@dp.callback_query_handler(lambda c: c.data.startswith("piece_"), state=OrderState.pieces)
async def process_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("pieces", [])
    piece = call.data.replace("piece_", "")
    if piece in selected:
        selected.remove(piece)
    else:
        selected.append(piece)
    await state.update_data(pieces=selected)
    await call.message.edit_reply_markup(reply_markup=get_pieces_kb(selected))

@dp.callback_query_handler(lambda c: c.data == "done_pieces", state=OrderState.pieces)
async def process_done_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    pieces = data.get("pieces", [])
    if not pieces:
        await call.answer("❌ اختر قطعة واحدة على الأقل!", show_alert=True)
        return
    need_over = any(p in pieces for p in ["أوفر", "سيت 3", "سيت 6"])
    need_hand = any(p in pieces for p in ["ملحف", "سيت 3", "سيت 6"])
    need_box = "بوكس ككو" in pieces
    need_dist = "توزيعات" in pieces
    await state.update_data(need_over=need_over, need_hand=need_hand, need_box=need_box, need_dist=need_dist)
    if need_over:
        await call.message.answer("✨ نوع الأوفر:", reply_markup=get_over_type_kb())
        await OrderState.over_type.set()
        return
    if need_hand:
        await call.message.answer("🛏 نوع الملحف:", reply_markup=get_hand_type_kb())
        await OrderState.hand_type.set()
        return
    if need_box:
        await call.message.answer("🎁 اختر لون البوكس:", reply_markup=get_box_color_kb())
        await OrderState.box_color.set()
        return
    if need_dist:
        await call.message.answer("🎉 اكتب عدد التوزيعات:")
        await OrderState.dist_count.set()
        return
    await call.message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

@dp.callback_query_handler(lambda c: c.data.startswith("over_"), state=OrderState.over_type)
async def process_over_type(call: types.CallbackQuery, state: FSMContext):
    over_choice = call.data.replace("over_", "")
    await state.update_data(over_type=over_choice)
    data = await state.get_data()
    if data.get("need_hand"):
        await call.message.answer("🛏 نوع الملحف:", reply_markup=get_hand_type_kb())
        await OrderState.hand_type.set()
        return
    if data.get("need_box"):
        await call.message.answer("🎁 اختر لون البوكس:", reply_markup=get_box_color_kb())
        await OrderState.box_color.set()
        return
    if data.get("need_dist"):
        await call.message.answer("🎉 اكتب عدد التوزيعات:")
        await OrderState.dist_count.set()
        return
    await call.message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

@dp.callback_query_handler(lambda c: c.data.startswith("hand_"), state=OrderState.hand_type)
async def process_hand_type(call: types.CallbackQuery, state: FSMContext):
    hand_choice = call.data.replace("hand_", "")
    await state.update_data(hand_type=hand_choice)
    data = await state.get_data()
    if data.get("need_box"):
        await call.message.answer("🎁 اختر لون البوكس:", reply_markup=get_box_color_kb())
        await OrderState.box_color.set()
        return
    if data.get("need_dist"):
        await call.message.answer("🎉 اكتب عدد التوزيعات:")
        await OrderState.dist_count.set()
        return
    await call.message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

@dp.callback_query_handler(lambda c: c.data.startswith("box_"), state=OrderState.box_color)
async def process_box_color(call: types.CallbackQuery, state: FSMContext):
    box_color = call.data.replace("box_", "")
    await state.update_data(box_color=box_color)
    data = await state.get_data()
    if data.get("need_dist"):
        await call.message.answer("🎉 اكتب عدد التوزيعات:")
        await OrderState.dist_count.set()
        return
    await call.message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

@dp.message_handler(state=OrderState.dist_count)
async def process_dist_count(msg: types.Message, state: FSMContext):
    count = msg.text.strip()
    try:
        if int(count) <= 0:
            raise ValueError
    except:
        await msg.answer("❌ أدخل رقماً صحيحاً أكبر من 0:")
        return
    await state.update_data(dist_count=count)
    await msg.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

@dp.callback_query_handler(lambda c: c.data.startswith("size_"), state=OrderState.size)
async def process_size(call: types.CallbackQuery, state: FSMContext):
    size = call.data.replace("size_", "")
    await state.update_data(size=size)
    await call.message.answer("💰 اكتب سعر الطلب:")
    await OrderState.price.set()

@dp.message_handler(state=OrderState.price)
async def process_price(msg: types.Message, state: FSMContext):
    price = msg.text.strip()
    if not validate_price(price):
        await msg.answer("❌ أدخل سعراً صحيحاً:")
        return
    await state.update_data(price=price)
    await msg.answer("📝 ملاحظات؟ (اكتب 'لا' بدون):")
    await OrderState.notes.set()

@dp.message_handler(state=OrderState.notes)
async def process_notes(msg: types.Message, state: FSMContext):
    notes = "لا يوجد" if msg.text.strip().lower() in ["لا", "لايوجد"] else msg.text.strip()
    await state.update_data(notes=notes)
    await msg.answer("📸 ارسل الصور (1-4) أو اكتب 'تم':")
    await state.update_data(images=[])
    await OrderState.images.set()

@dp.message_handler(content_types=['photo'], state=OrderState.images)
async def process_photo(msg: types.Message, state: FSMContext):
    data = await state.get_data()
    images = data.get("images", [])
    if len(images) >= 4:
        await msg.answer("❌ الحد الأقصى 4 صور!")
        return
    images.append(msg.photo[-1].file_id)
    await state.update_data(images=images)
    await msg.answer(f"✅ صورة ({len(images)}/4)")

@dp.message_handler(state=OrderState.images)
async def finish_order(msg: types.Message, state: FSMContext):
    if "تم" not in msg.text.lower():
        await msg.answer("❌ اكتب 'تم' أو أرسل صورة:")
        return
    
    try:
        order_id = get_next_order_id()
        data = await state.get_data()
        images_list = data.get("images", [])

        data["id"] = order_id

        orders_data[order_id] = {
            "data": data,
            "images": images_list,
            "current_group": "new"
        }

        save_to_excel(data, "orders.xlsx")

        text = format_order_text(data, order_id, "new")
        status_kb = get_status_buttons(order_id, "new")

        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            await bot.send_media_group(chat_id=GROUP_NEW, media=media)
        
        await bot.send_message(
            chat_id=GROUP_NEW, 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown'
        )
        
        await msg.answer(f"✅ طلب #{order_id} تم!")
    except Exception as e:
        print(f"❌ خطأ: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")
    finally:
        await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith("move_"))
async def move_order(call: types.CallbackQuery):
    try:
        parts = call.data.split("_")
        order_id = int(parts[1])
        target_group_name = parts[2]

        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return

        order_info = orders_data[order_id]
        data = order_info["data"]
        images_list = order_info["images"]
        current_group = order_info["current_group"]

        if current_group == target_group_name:
            await call.answer("🔔 موجود هنا!", show_alert=True)
            return

        target_group_id = GROUPS_MAP.get(target_group_name)
        text = format_order_text(data, order_id, target_group_name)
        status_kb = get_status_buttons(order_id, target_group_name)

        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            await bot.send_media_group(chat_id=target_group_id, media=media)
        
        await bot.send_message(
            chat_id=target_group_id, 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown'
        )

        # ✅ حذف الطلب مع الصور من الكروب السابق
        try:
            await call.message.delete()
            print(f"✅ تم حذف الرسالة والصور من الكروب السابق")
        except Exception as e:
            print(f"⚠️ لم يتمكن من حذف الرسالة: {e}")
        
        orders_data[order_id]["current_group"] = target_group_name

        target_name = GROUPS_NAMES.get(target_group_id)
        await call.answer(f"✅ {target_name}", show_alert=False)

    except Exception as e:
        print(f"❌ خطأ: {e}")
        await call.answer(f"❌ خطأ!", show_alert=True)

if __name__ == "__main__":
    print("🚀 البوت يعمل...")
    init_excel_file("orders.xlsx")
    executor.start_polling(dp, skip_updates=True)
